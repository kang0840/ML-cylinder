"""진동·소리·잔여수명 PyCaret 모델 통합 추론기.

JSON 한 건 또는 Excel의 1초 데이터 전체를 입력할 수 있다.
Excel 입력은 같은 파일의 ``추론결과`` 시트를 생성하거나 갱신한다.
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
import sys
from typing import Any

# PyCaret 3.x는 Python 3.12를 지원하지 않는다. 사용자가 일반 ``python``으로
# 실행해도 프로젝트의 Python 3.11 가상환경으로 자동 전환한다.
ROOT = Path(__file__).resolve().parent
if sys.version_info >= (3, 12):
    venv_python = ROOT / ".venv-pycaret" / "Scripts" / "python.exe"
    if venv_python.exists() and Path(sys.executable).resolve() != venv_python.resolve():
        os.execv(str(venv_python), [str(venv_python), str(Path(__file__).resolve()), *sys.argv[1:]])
    raise RuntimeError(
        "PyCaret은 Python 3.12를 지원하지 않습니다. "
        ".venv-pycaret(Python 3.11) 환경을 만들거나 그 Python으로 실행하세요."
    )

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill


FEATURES = [
    "mean", "standard_deviation", "rms", "maximum", "minimum", "peak",
    "peak_to_peak", "crest_factor", "dominant_frequency", "dominant_amplitude",
    "spectral_energy",
]
MODEL_PATHS = {
    "vibration": ROOT / "models/pycaret_final/pycaret_final_vibration_model.pkl",
    "sound": ROOT / "models/pycaret_final/pycaret_final_sound_model.pkl",
    "life": ROOT / "models/pycaret_final/pycaret_final_life_model.pkl",
}

# 잔여수명 모델은 기존 임계값 분류기가 만든 아래 세 신뢰도 값으로만
# 학습되었다. PyCaret 분류기의 0.9~1.0 확률을 직접 전달하면 학습 범위를
# 벗어나므로, 상태를 기존 학습 계약에 맞는 값으로 변환해서 전달한다.
LIFE_CONFIDENCE_BY_STATUS = {
    "normal": 0.55,
    "pressure_drop": 0.60,
    "seal_leak": 0.60,
    "internal_wear": 0.65,
    "unknown": 0.65,
}
STATUS_SEVERITY = {
    "normal": 0,
    "pressure_drop": 1,
    "seal_leak": 2,
    "internal_wear": 3,
    "unknown": 4,
}


class ThreeModelPredictor:
    def __init__(self) -> None:
        missing = [str(path) for path in MODEL_PATHS.values() if not path.exists()]
        if missing:
            raise FileNotFoundError("모델 파일을 찾을 수 없습니다: " + ", ".join(missing))
        self.packages = {name: joblib.load(path) for name, path in MODEL_PATHS.items()}

    @staticmethod
    def _frame(package: dict[str, Any], values: dict[str, Any]) -> pd.DataFrame:
        features = package["feature_names"]
        missing = [name for name in features if name not in values]
        if missing:
            raise ValueError(f"필수 특징값 누락: {', '.join(missing)}")
        return pd.DataFrame([{name: values[name] for name in features}], columns=features)

    @staticmethod
    def _condition_batch(package: dict[str, Any], frame: pd.DataFrame) -> tuple[np.ndarray, np.ndarray]:
        model = package["model"]
        predictions = np.asarray(model.predict(frame)).astype(str)
        if hasattr(model, "predict_proba"):
            confidence = np.max(np.asarray(model.predict_proba(frame), dtype=float), axis=1)
        else:
            confidence = np.zeros(len(frame), dtype=float)
        return predictions, confidence

    @staticmethod
    def _overall_status(vibration: str, sound: str) -> str:
        """두 센서 중 더 위험한 실제 분류 결과를 종합 상태로 사용한다."""
        return max((vibration, sound), key=lambda value: STATUS_SEVERITY.get(value, 4))

    def predict(self, vibration: dict[str, Any], sound: dict[str, Any]) -> dict[str, Any]:
        frames = {
            "vibration": self._frame(self.packages["vibration"], vibration),
            "sound": self._frame(self.packages["sound"], sound),
        }
        vp, vc = self._condition_batch(self.packages["vibration"], frames["vibration"])
        sp, sc = self._condition_batch(self.packages["sound"], frames["sound"])
        life_frame = pd.DataFrame([{
            "vibration_prediction": vp[0],
            "vibration_confidence": LIFE_CONFIDENCE_BY_STATUS.get(vp[0], 0.65),
            "sound_prediction": sp[0],
            "sound_confidence": LIFE_CONFIDENCE_BY_STATUS.get(sp[0], 0.65),
        }])
        remaining = float(self.packages["life"]["model"].predict(life_frame)[0])
        remaining = round(float(np.clip(remaining, 0, 100)), 2)
        return {
            "overall_status": self._overall_status(vp[0], sp[0]),
            "remaining_life_percent": remaining,
            "vibration": {"prediction": vp[0], "confidence": round(float(vc[0]), 6),
                          "model_version": self.packages["vibration"]["model_version"]},
            "sound": {"prediction": sp[0], "confidence": round(float(sc[0]), 6),
                      "model_version": self.packages["sound"]["model_version"]},
            "life_model_version": self.packages["life"]["model_version"],
        }

    def predict_excel(self, path: Path) -> tuple[int, dict[str, Any]]:
        source = pd.read_excel(path, sheet_name="1초 입력데이터")
        expected = ["초 번호"] + [f"진동_{x}" for x in FEATURES] + [f"소리_{x}" for x in FEATURES]
        missing = [column for column in expected if column not in source.columns]
        if missing:
            raise ValueError("Excel 필수 열 누락: " + ", ".join(missing))
        vibration = source[[f"진동_{x}" for x in FEATURES]].copy()
        vibration.columns = FEATURES
        sound = source[[f"소리_{x}" for x in FEATURES]].copy()
        sound.columns = FEATURES
        if vibration.isnull().any().any() or sound.isnull().any().any():
            raise ValueError("입력 데이터에 빈 셀이 있습니다.")

        vp, vc = self._condition_batch(self.packages["vibration"], vibration)
        sp, sc = self._condition_batch(self.packages["sound"], sound)
        life_frame = pd.DataFrame({
            "vibration_prediction": vp,
            "vibration_confidence": [LIFE_CONFIDENCE_BY_STATUS.get(value, 0.65) for value in vp],
            "sound_prediction": sp,
            "sound_confidence": [LIFE_CONFIDENCE_BY_STATUS.get(value, 0.65) for value in sp],
        })
        remaining = np.clip(self.packages["life"]["model"].predict(life_frame), 0, 100)
        statuses = [self._overall_status(vp[i], sp[i]) for i in range(len(source))]

        workbook = load_workbook(path)
        if "추론결과" in workbook.sheetnames:
            del workbook["추론결과"]
        sheet = workbook.create_sheet("추론결과")
        headers = ["초 번호", "진동 판정", "진동 신뢰도", "소리 판정", "소리 신뢰도",
                   "예상 잔여수명(%)", "종합 상태", "진동 모델 버전", "소리 모델 버전", "수명 모델 버전"]
        sheet.append(headers)
        versions = {name: self.packages[name]["model_version"] for name in self.packages}
        for index in range(len(source)):
            sheet.append([
                source.iloc[index]["초 번호"], vp[index], float(vc[index]), sp[index], float(sc[index]),
                round(float(remaining[index]), 2), statuses[index], versions["vibration"],
                versions["sound"], versions["life"],
            ])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        widths = [10, 16, 14, 16, 14, 18, 16, 37, 37, 37]
        for column, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + column)].width = width
        for column in ("C", "E"):
            for cell in sheet[column][1:]:
                cell.number_format = "0.0%"
        for cell in sheet["F"][1:]:
            cell.number_format = "0.0"
        end = sheet.max_row
        sheet.conditional_formatting.add(
            f"G2:G{end}", FormulaRule(formula=['G2="normal"'], fill=PatternFill("solid", fgColor="C6EFCE")))
        sheet.conditional_formatting.add(
            f"G2:G{end}", FormulaRule(formula=['G2<>"normal"'], fill=PatternFill("solid", fgColor="FFC7CE")))
        workbook.save(path)
        last = len(source) - 1
        representative = {
            "overall_status": statuses[last],
            "remaining_life_percent": round(float(remaining[last]), 2),
            "vibration": {
                "prediction": vp[last],
                "confidence": round(float(vc[last]), 6),
                "model_version": versions["vibration"],
            },
            "sound": {
                "prediction": sp[last],
                "confidence": round(float(sc[last]), 6),
                "model_version": versions["sound"],
            },
            "life_model_version": versions["life"],
            "second_number": int(source.iloc[last]["초 번호"]),
        }
        return len(source), representative


def print_result(result: dict[str, Any], input_name: str, total_rows: int | None = None) -> None:
    labels = {"normal": "정상", "pressure_drop": "압력 저하", "seal_leak": "점검 필요",
              "internal_wear": "위험", "unknown": "판정 불가"}
    print("\n" + "=" * 54)
    print("          스마트 실린더 AI 상태 진단")
    print("=" * 54)
    print(f"  종합 상태       : {labels.get(result['overall_status'], result['overall_status'])}")
    print(f"  예상 잔여수명   : {result['remaining_life_percent']:5.1f}%")
    print("-" * 54)
    print(f"  진동 센서       : {labels.get(result['vibration']['prediction'], result['vibration']['prediction'])}")
    print(f"  진동 신뢰도     : {result['vibration']['confidence'] * 100:5.1f}%")
    print(f"  소리 센서       : {labels.get(result['sound']['prediction'], result['sound']['prediction'])}")
    print(f"  소리 신뢰도     : {result['sound']['confidence'] * 100:5.1f}%")
    print("=" * 54)
    print(f"  입력 파일       : {input_name}\n")
    if total_rows is not None:
        print(f"  화면 표시 기준  : 마지막 {result['second_number']}초 데이터")
        print(f"  전체 처리 건수  : {total_rows:,}개 1초 데이터")
        print(f"  전체 결과 위치  : 같은 Excel의 '추론결과' 시트")
        print()


def main() -> None:
    parser = argparse.ArgumentParser(description="진동·소리·수명 PyCaret 모델 통합 추론")
    parser.add_argument("--input", type=str, default=str(ROOT / "inputs/three_model_input_example.json"),
                        help="JSON 또는 Excel(.xlsx) 파일 경로")
    parser.add_argument("--output", type=Path, help="JSON 결과 저장 위치(JSON 입력에서만 사용)")
    parser.add_argument("--json", action="store_true", help="JSON 입력 결과를 JSON으로 화면 출력")
    args = parser.parse_args()
    # PowerShell에서 복사한 경로 앞뒤에 공백이나 따옴표가 섞여도 정리한다.
    cleaned_input = args.input.strip().strip('"').strip("'").strip()
    input_path = Path(cleaned_input).expanduser().resolve()
    if not input_path.exists():
        parser.error(f"입력 파일을 찾을 수 없습니다: {input_path}")
    predictor = ThreeModelPredictor()
    if input_path.suffix.lower() == ".xlsx":
        count, representative = predictor.predict_excel(input_path)
        print_result(representative, input_path.name, count)
        return
    if input_path.suffix.lower() != ".json":
        parser.error("지원 파일 형식은 .json 또는 .xlsx입니다.")
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    result = predictor.predict(payload["vibration"], payload["sound"])
    rendered = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        args.output.write_text(rendered + "\n", encoding="utf-8")
        print(args.output.resolve())
    elif args.json:
        print(rendered)
    else:
        print_result(result, input_path.name)


if __name__ == "__main__":
    main()
