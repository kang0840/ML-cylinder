from openpyxl import load_workbook

from src.inference_excel_store import (
    AsyncInferenceExcelStore,
    HEADERS,
    InferenceExcelStore,
    SHEET_NAME,
)


def test_inference_is_written_read_back_and_marked_saved(tmp_path):
    path = tmp_path / "inference_results.xlsx"
    store = InferenceExcelStore(path)
    result = store.record_pending(
        "measurement-1", "2026-08-08T00:00:00+00:00", "pico-test",
        "sph0645", "forward",
        {"prediction": "normal", "confidence": 0.75,
         "health_score": 92.5, "model_version": "test-v1"},
    )
    assert result == {
        "prediction": "normal", "confidence": 0.75,
        "health_score": 92.5, "model_version": "test-v1",
    }
    store.mark_transferred("measurement-1")

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook[SHEET_NAME].iter_rows(values_only=True))
    assert rows[0] == HEADERS
    status=HEADERS.index("db_status")
    assert rows[1][status] == "saved"
    assert rows[1][status+1] is not None


def test_async_excel_store_batches_results_after_db_transfer(tmp_path):
    path = tmp_path / "async_results.xlsx"
    store = AsyncInferenceExcelStore(path, flush_seconds=0.01, batch_size=100)
    for index in range(3):
        measurement_id = f"measurement-{index}"
        store.record_pending(
            measurement_id, "2026-08-08T00:00:00+00:00", "pico-test",
            "sph0645", "forward",
            {"prediction": "normal", "confidence": 0.75,
             "health_score": 92.5, "model_version": "test-v1"},
        )
        store.mark_transferred(measurement_id)
    store.close()

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook[SHEET_NAME].iter_rows(values_only=True))
    assert len(rows) == 4
    status=HEADERS.index("db_status")
    assert all(row[status] == "saved" for row in rows[1:])
