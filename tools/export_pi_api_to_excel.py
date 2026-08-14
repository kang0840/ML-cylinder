from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.error import URLError
from urllib.request import urlopen
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


API_URL = "http://192.168.137.99:8000/api/real-cylinder"
OUTPUT = Path("outputs/pi5_inference/pi5_inference_results.xlsx")
KST = ZoneInfo("Asia/Seoul")


def parse_time(value: str) -> datetime:
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(KST).replace(tzinfo=None)


try:
    with urlopen(API_URL, timeout=5) as response:
        payload = json.load(response)
except (URLError, TimeoutError):
    from dotenv import load_dotenv
    from supabase import create_client

    load_dotenv()
    import os

    url = os.environ["SUPABASE_URL"]
    key = os.environ["SUPABASE_KEY"]
    table = os.environ.get("SUPABASE_TABLE", "smart_cylinder_analysis")
    result = (
        create_client(url, key)
        .table(table)
        .select("*")
        .order("measured_at", desc=True)
        .limit(120)
        .execute()
    )
    rows = list(reversed(result.data or []))
    for index, row in enumerate(rows, 1):
        row.setdefault("sequence", index)
        row.setdefault("controlling_role", "-")
        row.setdefault("cylinder_state", "-")
        row.setdefault("fusion_version", row.get("model_version", "-"))
        row.setdefault("vibration_rms", None)
        row.setdefault("sound_rms", None)
    payload = {
        "history": rows,
        "latest": rows[-1] if rows else {},
        "source": "supabase_pi_mirror",
    }

history = payload.get("history", [])
latest = payload.get("latest") or (history[-1] if history else {})
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
summary = wb.active
summary.title = "요약"
data = wb.create_sheet("추론결과")

summary.append(["Raspberry Pi 5 추론 데이터 요약"])
summary.merge_cells("A1:D1")
summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
summary["A1"].alignment = Alignment(horizontal="center")
summary.append(["수집 API", API_URL])
summary.append(["Excel 생성 시각(KST)", datetime.now(KST).replace(tzinfo=None)])
summary.append(["데이터 건수", len(history)])
summary.append(["최신 측정 시각(KST)", parse_time(latest["measured_at"]) if latest else None])
summary.append(["최신 판정", latest.get("prediction", "-")])
summary.append(["최신 건강 점수", latest.get("health_score")])
summary.append(["최신 신뢰도", latest.get("confidence")])
summary.append(["데이터 출처", payload.get("source", "-")])
summary["B3"].number_format = "yyyy-mm-dd hh:mm:ss"
summary["B5"].number_format = "yyyy-mm-dd hh:mm:ss"
summary["B8"].number_format = "0.0%"
for row in range(2, 10):
    summary.cell(row, 1).font = Font(bold=True)
    summary.cell(row, 1).fill = PatternFill("solid", fgColor="D9EAF7")
summary.column_dimensions["A"].width = 23
summary.column_dimensions["B"].width = 48
summary.sheet_view.showGridLines = False

headers = [
    "측정 시각(KST)", "시퀀스", "판정", "판정(한글)", "건강 점수", "신뢰도",
    "제어 센서", "실린더 상태", "진동 RMS", "소리 RMS", "융합 모델 버전",
]
data.append(headers)
translations = {
    "normal": "정상", "pressure_drop": "압력 저하", "seal_leak": "실링 누설",
    "internal_wear": "내부 마모", "unknown": "알 수 없음",
}
for item in history:
    prediction = item.get("prediction", "unknown")
    data.append([
        parse_time(item["measured_at"]), item.get("sequence"), prediction,
        translations.get(prediction, prediction), item.get("health_score"),
        item.get("confidence"), item.get("controlling_role"), item.get("cylinder_state"),
        item.get("vibration_rms"), item.get("sound_rms"), item.get("fusion_version"),
    ])

header_fill = PatternFill("solid", fgColor="1F4E78")
for cell in data[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
data.freeze_panes = "A2"
data.auto_filter.ref = data.dimensions
for cell in data["A"][1:]:
    cell.number_format = "yyyy-mm-dd hh:mm:ss"
for cell in data["E"][1:]:
    cell.number_format = "0.0"
for cell in data["F"][1:]:
    cell.number_format = "0.0%"
for col in ("I", "J"):
    for cell in data[col][1:]:
        cell.number_format = "#,##0.00"
data.conditional_formatting.add(
    f"C2:C{data.max_row}", FormulaRule(formula=["C2=\"normal\""], fill=PatternFill("solid", fgColor="C6EFCE"))
)
data.conditional_formatting.add(
    f"C2:C{data.max_row}", FormulaRule(formula=["C2<>\"normal\""], fill=PatternFill("solid", fgColor="FFC7CE"))
)
widths = [21, 10, 18, 14, 12, 10, 14, 14, 16, 16, 38]
for index, width in enumerate(widths, 1):
    data.column_dimensions[get_column_letter(index)].width = width
data.sheet_view.showGridLines = False

wb.save(OUTPUT)
checked = load_workbook(OUTPUT, read_only=True, data_only=False)
assert checked.sheetnames == ["요약", "추론결과"]
assert checked["추론결과"].max_row == len(history) + 1
print(f"output={OUTPUT.resolve()}")
print(f"rows={len(history)}")
print(f"latest={latest.get('prediction', '-') if latest else '-'}")
