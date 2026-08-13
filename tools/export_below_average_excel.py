from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs/pi5_inference/below_average_model_input.xlsx"
FEATURES = [
    "mean", "standard_deviation", "rms", "maximum", "minimum", "peak",
    "peak_to_peak", "crest_factor", "dominant_frequency", "dominant_amplitude",
    "spectral_energy",
]


def load_sensor(name: str) -> pd.DataFrame:
    base = ROOT / "data/export/pycaret_server"
    return pd.concat(
        [pd.read_csv(base / f"{name}_train.csv"), pd.read_csv(base / f"{name}_test.csv")],
        ignore_index=True,
    )


def lower(value: float, ratio: float = 0.8) -> float:
    # 모든 값이 원래 값보다 숫자상 낮아지도록 처리한다.
    value = float(value)
    return value * ratio if value >= 0 else value - abs(value) * (1 - ratio)


vibration = load_sensor("vibration")
sound = load_sensor("sound")
count = min(len(vibration), len(sound))

wb = Workbook()
ws = wb.active
ws.title = "1초 입력데이터"

headers = ["초 번호"]
headers += [f"진동_{feature}" for feature in FEATURES]
headers += [f"소리_{feature}" for feature in FEATURES]
ws.append(headers)

for index in range(count):
    row = [index + 1]
    row += [lower(vibration.at[index, feature]) for feature in FEATURES]
    row += [lower(sound.at[index, feature]) for feature in FEATURES]
    ws.append(row)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 38
ws.freeze_panes = "B2"
ws.auto_filter.ref = ws.dimensions
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 10
for column in range(2, len(headers) + 1):
    ws.column_dimensions[get_column_letter(column)].width = 23
    for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=ws.max_row):
        for item in cell:
            item.number_format = "#,##0.0000"

guide = wb.create_sheet("사용방법")
guide.append(["항목", "내용"])
guide.append(["데이터 구조", "한 행이 1초분 모델 입력이며 진동 11개와 소리 11개 특징을 포함합니다."])
guide.append(["행 개수", count])
guide.append(["값 조정", "각 학습 데이터 행의 값을 원래 값보다 20% 낮게 조정했습니다."])
guide.append(["주의", "실제 측정 시각이 없는 학습 데이터이므로 초 번호는 순차 번호입니다."])
guide.append(["모델 입력", "통합 실행 파일이 이 Excel의 각 행을 순서대로 읽도록 연결해야 합니다."])
for cell in guide[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
guide.column_dimensions["A"].width = 18
guide.column_dimensions["B"].width = 85
guide.sheet_view.showGridLines = False

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
check = load_workbook(OUTPUT, read_only=True, data_only=True)
assert check["1초 입력데이터"].max_row == count + 1
assert check["1초 입력데이터"].max_column == 23
print(OUTPUT.resolve())
print(f"one_second_rows={count}")
