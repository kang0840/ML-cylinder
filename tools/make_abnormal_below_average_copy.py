"""1초 Excel 입력을 복사하고 모든 특징을 해당 열 평균보다 낮게 변환한다."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "outputs/pi5_inference/below_average_model_input.xlsx"
OUTPUT = ROOT / "outputs/pi5_inference/abnormal_below_average_test.xlsx"
SHEET = "1초 입력데이터"


if not SOURCE.exists():
    raise FileNotFoundError(f"기준 Excel을 찾을 수 없습니다: {SOURCE}")

shutil.copy2(SOURCE, OUTPUT)
workbook = load_workbook(OUTPUT)
sheet = workbook[SHEET]

# B:W는 진동 11개 + 소리 11개 특징이다. 초 번호(A)는 변경하지 않는다.
for column in range(2, sheet.max_column + 1):
    values = [float(sheet.cell(row, column).value) for row in range(2, sheet.max_row + 1)]
    average = sum(values) / len(values)
    scale = max(abs(average), max(abs(value) for value in values), 1.0)
    epsilon = scale * 1e-6
    for row, original in enumerate(values, 2):
        # 원래 값의 상대적 차이는 일부 유지하되 모든 결과를 평균 미만으로 제한한다.
        candidate = average - abs(original - average) * 0.8 - epsilon
        sheet.cell(row, column).value = min(candidate, average - epsilon)

if "추론결과" in workbook.sheetnames:
    del workbook["추론결과"]

if "사용방법" in workbook.sheetnames:
    guide = workbook["사용방법"]
    guide.append(["비정상 시험 복사본", "각 특징 열의 모든 1초 값을 해당 열 평균보다 낮게 변환한 민감도 시험 데이터입니다."])
    guide.append(["해석 주의", "평균 미만은 실제 고장 라벨이 아니므로 비정상 판정을 보장하지 않습니다."])

workbook.save(OUTPUT)

# 모든 특징값이 변환 전 기준 평균보다 낮은지 확인한다.
check = load_workbook(OUTPUT, read_only=True, data_only=True)[SHEET]
source_check = load_workbook(SOURCE, read_only=True, data_only=True)[SHEET]
transformed_rows = list(check.iter_rows(min_row=2, values_only=True))
source_rows = list(source_check.iter_rows(min_row=2, values_only=True))
for column in range(1, check.max_column):
    transformed = [float(row[column]) for row in transformed_rows]
    original_average = sum(float(row[column]) for row in source_rows) / len(source_rows)
    assert max(transformed) < original_average

print(OUTPUT.resolve())
print(f"1초 데이터 행 수={check.max_row - 1}")
