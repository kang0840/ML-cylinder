from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client


OUTPUT = Path("outputs/pi5_inference/pi5_one_second_measurements.xlsx")
KST = ZoneInfo("Asia/Seoul")


def kst(value: str | None):
    if not value:
        return None
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(KST).replace(tzinfo=None)


load_dotenv()
client = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])
table = os.environ.get("SUPABASE_TABLE", "smart_cylinder_analysis")

# Supabase의 기본 응답 제한을 피하면서 전체 측정값을 페이지 단위로 수집한다.
rows = []
page_size = 1000
for start in range(0, 100_000, page_size):
    page = (
        client.table(table).select("*").order("measured_at").range(start, start + page_size - 1).execute().data
        or []
    )
    rows.extend(page)
    if len(page) < page_size:
        break

grouped = defaultdict(dict)
for row in rows:
    grouped[row["measured_at"]][row.get("sensor_type", "unknown")] = row

wb = Workbook()
ws = wb.active
ws.title = "1초 측정데이터"
detail = wb.create_sheet("센서별 원본")

headers = [
    "측정 시각(KST)", "순번", "장치 ID", "실린더 상태",
    "진동 RMS", "진동 주요주파수(Hz)", "진동 주요진폭", "진동 판정", "진동 건강점수", "진동 신뢰도", "진동 모델",
    "소리 RMS", "소리 주요주파수(Hz)", "소리 주요진폭", "소리 판정", "소리 건강점수", "소리 신뢰도", "소리 모델",
]
ws.append(headers)
for sequence, measured_at in enumerate(sorted(grouped), 1):
    sensors = grouped[measured_at]
    vibration = sensors.get("sph0645", {})
    sound = sensors.get("inmp441", {})
    common = vibration or sound
    ws.append([
        kst(measured_at), sequence, common.get("device_id"), common.get("cylinder_state"),
        vibration.get("vibration_rms"), vibration.get("dominant_frequency"), vibration.get("dominant_amplitude"),
        vibration.get("prediction"), vibration.get("health_score"), vibration.get("confidence"), vibration.get("model_version"),
        sound.get("sound_rms"), sound.get("dominant_frequency"), sound.get("dominant_amplitude"),
        sound.get("prediction"), sound.get("health_score"), sound.get("confidence"), sound.get("model_version"),
    ])

detail_headers = [
    "measurement_id", "device_id", "sensor_type", "measured_at_KST", "cylinder_state",
    "vibration_rms", "sound_rms", "dominant_frequency", "dominant_amplitude",
    "prediction", "confidence", "health_score", "model_version", "created_at_KST",
]
detail.append(detail_headers)
for row in rows:
    detail.append([
        row.get("measurement_id"), row.get("device_id"), row.get("sensor_type"), kst(row.get("measured_at")),
        row.get("cylinder_state"), row.get("vibration_rms"), row.get("sound_rms"),
        row.get("dominant_frequency"), row.get("dominant_amplitude"), row.get("prediction"),
        row.get("confidence"), row.get("health_score"), row.get("model_version"), kst(row.get("created_at")),
    ])

for sheet in (ws, detail):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

for cell in ws["A"][1:]:
    cell.number_format = "yyyy-mm-dd hh:mm:ss"
for col in ("J", "Q"):
    for cell in ws[col][1:]:
        cell.number_format = "0.0%"
for col in ("A", "D", "N"):
    for cell in detail[col][1:]:
        if col in ("D", "N"):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
for cell in detail["K"][1:]:
    cell.number_format = "0.0%"

for status_col in ("H", "O"):
    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}{ws.max_row}",
        FormulaRule(formula=[f'{status_col}2="normal"'], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        f"{status_col}2:{status_col}{ws.max_row}",
        FormulaRule(formula=[f'AND({status_col}2<>"",{status_col}2<>"normal")'], fill=PatternFill("solid", fgColor="FFC7CE")),
    )

widths = [21, 9, 14, 14, 16, 19, 17, 16, 15, 13, 35, 16, 19, 17, 16, 15, 13, 35]
for index, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(index)].width = width
for index, width in enumerate([38, 14, 12, 21, 14, 16, 16, 20, 20, 16, 12, 13, 35, 21], 1):
    detail.column_dimensions[get_column_letter(index)].width = width

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
check = load_workbook(OUTPUT, read_only=True, data_only=True)
assert check["1초 측정데이터"].max_row == len(grouped) + 1
assert check["센서별 원본"].max_row == len(rows) + 1
print(f"output={OUTPUT.resolve()}")
print(f"one_second_rows={len(grouped)}")
print(f"sensor_rows={len(rows)}")
print(f"period={min(grouped) if grouped else '-'} .. {max(grouped) if grouped else '-'}")
