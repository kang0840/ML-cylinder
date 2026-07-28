"""Excel persistence and Random Forest training for cylinder features."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from sklearn.ensemble import RandomForestClassifier


EXCEL_FEATURES = (
    "rms_velocity_mm_s", "peak_velocity_mm_s", "crest_factor",
    "dominant_frequency_hz", "sound_rms", "operating_hours",
)
MODEL_FEATURES = EXCEL_FEATURES[:-1]
HEADERS = ("timestamp", *EXCEL_FEATURES, "label", "label_source")
SHEET_NAME = "TrainingData"


def create_excel_dataset(path: Path, x: np.ndarray, y: np.ndarray) -> None:
    """Create the initial workbook from temporary labelled training data."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    timestamp = dt.datetime.now(dt.timezone.utc).isoformat()
    for values, label in zip(np.asarray(x, dtype=float), np.asarray(y, dtype=int)):
        sheet.append((timestamp, *map(float, values), int(label), "v4_demo_seed"))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def append_measurement(path: Path, timestamp: str, features: dict[str, float],
                       label: int, label_source: str = "project_statistical_zone") -> None:
    """Append one sensor-feature record and its supervised-learning label."""
    if label not in range(4):
        raise ValueError("label must be 0, 1, 2, or 3")
    workbook = load_workbook(path)
    sheet = workbook[SHEET_NAME]
    sheet.append((timestamp, *(float(features[name]) for name in EXCEL_FEATURES),
                  int(label), label_source))
    workbook.save(path)


def load_rms_history(path: Path) -> list[tuple[float, float]]:
    """Load previously logged measurements for RUL trend continuity."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None or tuple(headers) != HEADERS:
        raise ValueError(f"Excel columns must be: {', '.join(HEADERS)}")
    rms_index = headers.index("rms_velocity_mm_s")
    hours_index = headers.index("operating_hours")
    source_index = headers.index("label_source")
    history = []
    for row in rows:
        if row[source_index] != "project_statistical_zone":
            continue
        try:
            hours, rms = float(row[hours_index]), float(row[rms_index])
        except (TypeError, ValueError):
            continue
        if np.isfinite(hours) and np.isfinite(rms):
            history.append((hours, rms))
    return history


def load_training_data(path: Path) -> tuple[np.ndarray, np.ndarray]:
    """Read and validate labelled feature rows from the workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None or tuple(headers) != HEADERS:
        raise ValueError(f"Excel columns must be: {', '.join(HEADERS)}")
    x, y = [], []
    for row_number, row in enumerate(rows, start=2):
        if all(value is None for value in row):
            continue
        try:
            values = [float(row[headers.index(name)]) for name in EXCEL_FEATURES]
            label = int(row[headers.index("label")])
        except (TypeError, ValueError) as error:
            raise ValueError(f"invalid Excel data at row {row_number}") from error
        if not np.all(np.isfinite(values)) or label not in range(4):
            raise ValueError(f"invalid Excel data at row {row_number}")
        x.append(values)
        y.append(label)
    if len(x) < 8 or len(set(y)) < 2:
        raise ValueError("Excel needs at least 8 labelled rows from two or more zones")
    return np.asarray(x), np.asarray(y)


def train_model_from_excel(path: Path, seed: int = 42) -> tuple[RandomForestClassifier, int]:
    """Train from sensor features and exclude automatically generated labels.

    Operating hours are deliberately excluded: they describe exposure, not the
    measured condition, and otherwise leak the demo label schedule into the
    classifier.  Statistical-zone rows are retained as history but are not fed
    back into the supervised model as if they were ground truth.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None or tuple(headers) != HEADERS:
        raise ValueError(f"Excel columns must be: {', '.join(HEADERS)}")
    x, y = [], []
    for row in rows:
        if row[headers.index("label_source")] == "project_statistical_zone":
            continue
        try:
            values = [float(row[headers.index(name)]) for name in MODEL_FEATURES]
            label = int(row[headers.index("label")])
        except (TypeError, ValueError) as error:
            raise ValueError("invalid trusted training data") from error
        if not np.all(np.isfinite(values)) or label not in range(4):
            raise ValueError("invalid trusted training data")
        x.append(values)
        y.append(label)
    if len(x) < 8 or len(set(y)) < 2:
        raise ValueError("Excel needs at least 8 trusted labelled rows from two or more zones")
    x_array, y_array = np.asarray(x), np.asarray(y)
    model = RandomForestClassifier(
        n_estimators=200, max_depth=8, min_samples_leaf=3,
        class_weight="balanced", random_state=seed, n_jobs=-1,
        bootstrap=True, oob_score=True,
    ).fit(x_array, y_array)
    return model, len(y_array)


def load_zone_d_rms_threshold(path: Path) -> float:
    """Return the C/D RMS boundary inferred from trusted labelled examples."""
    x, y = load_training_data(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[SHEET_NAME]
    sources = [row[0] for row in sheet.iter_rows(
        min_row=2, min_col=HEADERS.index("label_source") + 1,
        max_col=HEADERS.index("label_source") + 1, values_only=True
    )]
    trusted = np.asarray([source != "project_statistical_zone" for source in sources])
    x, y = x[trusted], y[trusted]
    c_rms, d_rms = x[y == 2, 0], x[y == 3, 0]
    if len(d_rms) == 0:
        raise ValueError("trusted training data needs Zone D examples for RUL")
    if len(c_rms):
        return float((np.median(c_rms) + np.median(d_rms)) / 2.0)
    lower = x[y < 3, 0]
    if len(lower) == 0:
        raise ValueError("trusted training data needs a zone below D for RUL")
    return float((np.median(lower) + np.median(d_rms)) / 2.0)
