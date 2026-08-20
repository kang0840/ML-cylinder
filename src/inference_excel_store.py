"""추론 결과와 실험 조건을 Excel에 비동기 저장한다."""
from __future__ import annotations
import queue,threading
from datetime import datetime,timezone
from pathlib import Path
from typing import Any
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAME="InferenceResults"
HEADERS=("measurement_id","timestamp","device_id","sensor_type","sensor_role","experiment_id","session_id","cylinder_state","pressure_mpa","load_kg","ground_truth","ground_truth_source","prediction","confidence","rms","peak","crest_factor","dominant_frequency","dominant_amplitude","health_score","model_version","db_status","db_saved_at","db_error")

def _row(mid,measured,device,sensor,state,result,metadata=None,features=None,status="pending",saved=None,error=None):
    m=metadata or {};f=features or {}
    return (mid,measured,device,sensor,m.get("sensor_role"),m.get("experiment_id"),m.get("session_id"),state,m.get("pressure_mpa"),m.get("load_kg"),m.get("ground_truth"),m.get("ground_truth_source"),result["prediction"],float(result["confidence"]),f.get("rms"),f.get("peak"),f.get("crest_factor"),f.get("dominant_frequency"),f.get("dominant_amplitude"),float(result["health_score"]),result["model_version"],status,saved,error)

class InferenceExcelStore:
    def __init__(self,path:Path):self.path=path;self.path.parent.mkdir(parents=True,exist_ok=True);self._lock=threading.RLock()
    def _create_workbook(self):
        wb=Workbook();ws=wb.active;ws.title=SHEET_NAME;ws.append(HEADERS);ws.freeze_panes="A2";ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}1";ws.sheet_view.showGridLines=False
        for cell in ws[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="1F4E78")
        for i,name in enumerate(HEADERS,1):ws.column_dimensions[get_column_letter(i)].width=max(12,min(32,len(name)+4))
        wb.save(self.path)
    def record_pending(self,measurement_id,measured_at,device_id,sensor_type,cylinder_state,result,metadata=None,features=None):
        with self._lock:
            if not self.path.exists():self._create_workbook()
            wb=load_workbook(self.path);ws=wb[SHEET_NAME];ws.append(_row(measurement_id,measured_at,device_id,sensor_type,cylinder_state,result,metadata,features));ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}";wb.save(self.path);wb.close()
        return {k:(float(result[k]) if k in {"confidence","health_score"} else str(result[k])) for k in ("prediction","confidence","health_score","model_version")}
    def mark_transferred(self,measurement_id,error=None):
        with self._lock:
            wb=load_workbook(self.path);ws=wb[SHEET_NAME];status_col=HEADERS.index("db_status")+1
            for n in range(ws.max_row,1,-1):
                if ws.cell(n,1).value==measurement_id:
                    ws.cell(n,status_col).value="failed" if error else "saved";ws.cell(n,status_col+1).value=None if error else datetime.now(timezone.utc).isoformat();ws.cell(n,status_col+2).value=error;wb.save(self.path);wb.close();return
            wb.close();raise KeyError(f"Excel 추론 행을 찾을 수 없습니다: {measurement_id}")
    def append_saved_batch(self,rows):
        if not rows:return
        with self._lock:
            if not self.path.exists():self._create_workbook()
            wb=load_workbook(self.path);ws=wb[SHEET_NAME]
            for values in rows:ws.append(values)
            ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}";wb.save(self.path);wb.close()

class AsyncInferenceExcelStore:
    def __init__(self,path:Path,flush_seconds:float=2.0,batch_size:int=100):
        self.store=InferenceExcelStore(path);self.flush_seconds=flush_seconds;self.batch_size=batch_size;self._pending={};self._lock=threading.RLock();self._queue=queue.Queue();self._thread=threading.Thread(target=self._run,daemon=True);self._thread.start()
    def record_pending(self,measurement_id,measured_at,device_id,sensor_type,cylinder_state,result,metadata=None,features=None):
        normalized={k:(float(result[k]) if k in {"confidence","health_score"} else str(result[k])) for k in ("prediction","confidence","health_score","model_version")}
        with self._lock:self._pending[measurement_id]=_row(measurement_id,measured_at,device_id,sensor_type,cylinder_state,normalized,metadata,features,"saved",datetime.now(timezone.utc).isoformat(),None)
        return normalized
    def mark_transferred(self,measurement_id,error=None):
        with self._lock:row=self._pending.pop(measurement_id,None)
        if row is None:return
        if error:row=(*row[:-3],"failed",None,error)
        self._queue.put(row)
    def _run(self):
        while True:
            first=self._queue.get()
            if first is None:return
            batch=[first];deadline=datetime.now().timestamp()+self.flush_seconds
            while len(batch)<self.batch_size:
                try:item=self._queue.get(timeout=max(0,deadline-datetime.now().timestamp()))
                except queue.Empty:break
                if item is None:self.store.append_saved_batch(batch);return
                batch.append(item)
            self.store.append_saved_batch(batch)
    def close(self):self._queue.put(None);self._thread.join(timeout=10)
